
import copy
import re

from api.db import ParserType
from io import BytesIO
from rag.nlp import rag_tokenizer, tokenize, tokenize_table, add_positions, bullets_category, title_frequency, tokenize_chunks, docx_question_level
from deepdoc.parser import PdfParser, PlainParser
from rag.utils import num_tokens_from_string
from deepdoc.parser import PdfParser, ExcelParser, DocxParser
from docx import Document
from PIL import Image
import openpyxl

def logger(prog=None, msg=""):
        print(msg)

class Pdf(PdfParser):
    def __init__(self):
        self.model_speciess = ParserType.MANUAL.value
        super().__init__()

    def __call__(self, filename, binary=None, from_page=0,
                 to_page=100000, zoomin=3, callback=None):
        
        from timeit import default_timer as timer
        start = timer()
        callback(msg="OCR is running...")
        self.__images__(
            filename if not binary else binary,
            zoomin,
            from_page,
            to_page,
            callback
        )
        callback(msg="OCR finished.")
        print("OCR:", timer() - start)

        self._layouts_rec(zoomin)
        callback(0.65, "Layout analysis finished.")
        print("layouts:", timer() - start)
        self._table_transformer_job(zoomin)
        callback(0.67, "Table analysis finished.")
        self._text_merge()
        tbls = self._extract_table_figure(True, zoomin, True, True)
        self._concat_downward()
        self._filter_forpages()
        callback(0.68, "Text merging finished")

        # clean mess
        for b in self.boxes:
            b["text"] = re.sub(r"([\t 　]|\u3000){2,}", " ", b["text"].strip())

        return [(b["text"], b.get("layout_no", ""), self.get_position(b, zoomin))
                for i, b in enumerate(self.boxes)], tbls

def chunk(filename, binary=None, from_page=0, to_page=100000,
          lang="Chinese", callback=None, **kwargs):
    """
        Only pdf is supported.
    """
    pdf_parser = None
    doc = {
        "docnm_kwd": filename
    }
    doc["title_tks"] = rag_tokenizer.tokenize(re.sub(r"\.[a-zA-Z]+$", "", doc["docnm_kwd"]))
    doc["title_sm_tks"] = rag_tokenizer.fine_grained_tokenize(doc["title_tks"])
    # is it English
    eng = lang.lower() == "english"  # pdf_parser.is_english
    if re.search(r"\.pdf$", filename, re.IGNORECASE):
        pdf_parser = Pdf() if kwargs.get(
            "parser_config", {}).get(
            "layout_recognize", True) else PlainParser()
        sections, tbls = pdf_parser(filename if not binary else binary,
                                    from_page=from_page, to_page=to_page, callback=callback)
        if sections and len(sections[0]) < 3:
            sections = [(t, l, [[0] * 5]) for t, l in sections]
        # set pivot using the most frequent type of title,
        # then merge between 2 pivot
        if len(sections) > 0 and len(pdf_parser.outlines) / len(sections) > 0.1:
            max_lvl = max([lvl for _, lvl in pdf_parser.outlines])
            most_level = max(0, max_lvl - 1)
            levels = []
            for txt, _, _ in sections:
                for t, lvl in pdf_parser.outlines:
                    tks = set([t[i] + t[i + 1] for i in range(len(t) - 1)])
                    tks_ = set([txt[i] + txt[i + 1]
                                for i in range(min(len(t), len(txt) - 1))])
                    if len(set(tks & tks_)) / max([len(tks), len(tks_), 1]) > 0.8:
                        levels.append(lvl)
                        break
                else:
                    levels.append(max_lvl + 1)

        else:
            bull = bullets_category([txt for txt, _, _ in sections])
            most_level, levels = title_frequency(
                bull, [(txt, l) for txt, l, poss in sections])

        assert len(sections) == len(levels)
        sec_ids = []
        sid = 0
        for i, lvl in enumerate(levels):
            if lvl <= most_level and i > 0 and lvl != levels[i - 1]:
                sid += 1
            sec_ids.append(sid)
            # print(lvl, self.boxes[i]["text"], most_level, sid)

        sections = [(txt, sec_ids[i], poss)
                    for i, (txt, _, poss) in enumerate(sections)]
        for (img, rows), poss in tbls:
            if not rows: continue
            sections.append((rows if isinstance(rows, str) else rows[0], -1,
                            [(p[0] + 1 - from_page, p[1], p[2], p[3], p[4]) for p in poss]))

        def tag(pn, left, right, top, bottom):
            if pn + left + right + top + bottom == 0:
                return ""
            return "@@{}\t{:.1f}\t{:.1f}\t{:.1f}\t{:.1f}##" \
                .format(pn, left, right, top, bottom)

        chunks = []
        last_sid = -2
        tk_cnt = 0
        for txt, sec_id, poss in sorted(sections, key=lambda x: (
                x[-1][0][0], x[-1][0][3], x[-1][0][1])):
            poss = "\t".join([tag(*pos) for pos in poss])
            if tk_cnt < 32 or (tk_cnt < 1024 and (sec_id == last_sid or sec_id == -1)):
                if chunks:
                    chunks[-1] += "\n" + txt + poss
                    tk_cnt += num_tokens_from_string(txt)
                    continue
            chunks.append(txt + poss)
            tk_cnt = num_tokens_from_string(txt)
            if sec_id > -1:
                last_sid = sec_id

        res = tokenize_table(tbls, doc, eng)
        res.extend(tokenize_chunks(chunks, doc, eng, pdf_parser))
        return res
    
def read_and_process_excel(file_path):
    
    data = []
    # 打开 Excel 文件
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    # 创建一个字典来存储单元格值
    cell_values = {}

    # 遍历所有单元格
    for row in sheet.iter_rows():
        for cell in row:
            cell_values[cell.coordinate] = cell.value

    # 遍历所有合并单元格
    for merged_range in sheet.merged_cells.ranges:
        min_col, min_row = merged_range.min_col, merged_range.min_row
        max_col, max_row = merged_range.max_col, merged_range.max_row
        
        # 获取合并单元格的值（通常位于合并区域的左上角）
        value = sheet.cell(row=min_row, column=min_col).value
        
        # 将值填充到所有合并区域的单元格中
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                cell_coordinate = openpyxl.utils.get_column_letter(col) + str(row)
                cell_values[cell_coordinate] = value

    # 输出所有单元格的值
    max_row = sheet.max_row
    max_col = sheet.max_column
    for row in range(1, max_row + 1):
        row_values = []
        for col in range(1, max_col + 1):
            cell_coordinate = openpyxl.utils.get_column_letter(col) + str(row)
            cell_value = cell_values.get(cell_coordinate, None)
            row_values.append(str(cell_value).replace('\n', '') if cell_value is not None else '')
        data.append(row_values)
    return data